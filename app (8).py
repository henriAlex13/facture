import streamlit as st
import pandas as pd
import io
import pickle
import os
from datetime import datetime

st.set_page_config(
    page_title="Gestion Factures - SOCODECI & CIE",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .main-title {
        color: white;
        font-size: 2.5rem;
        font-weight: bold;
        margin: 0;
        text-align: center;
    }
    .main-subtitle {
        color: #f0f0f0;
        font-size: 1.2rem;
        text-align: center;
        margin-top: 0.5rem;
    }
    .badge-coming {
        background: #e0e0e0;
        color: #666;
        padding: 0.3rem 1rem;
        border-radius: 20px;
        font-style: italic;
        display: inline-block;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8f9fa 0%, #e9ecef 100%);
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        text-align: center;
    }
    .stButton>button {
        border-radius: 8px;
        font-weight: 500;
    }
</style>
""", unsafe_allow_html=True)

# Fichiers Excel principaux
FICHIER_BT = "FACTURAT_ELECTRICITE_BT.xlsx"
FICHIER_HT = "FACTURAT_ELECTRICITE_HT.xlsx"

# Fichiers de sauvegarde (backup automatique)
SAVE_FILE_BT = "data_factures_cie_bt.pkl"
SAVE_FILE_HT = "data_factures_cie_ht.pkl"

# Fonctions de chargement
def load_data_bt():
    # Essayer d'abord le pickle (sauvegarde)
    if os.path.exists(SAVE_FILE_BT):
        with open(SAVE_FILE_BT, 'rb') as f:
            return pickle.load(f)
    # Sinon charger depuis Excel
    elif os.path.exists(FICHIER_BT):
        return pd.read_excel(FICHIER_BT)
    else:
        st.error(f"""
        ❌ **Fichier BT introuvable !**
        
        Veuillez placer le fichier `{FICHIER_BT}` dans le même dossier que l'application.
        """)
        st.stop()

def load_data_ht():
    # Essayer d'abord le pickle (sauvegarde)
    if os.path.exists(SAVE_FILE_HT):
        with open(SAVE_FILE_HT, 'rb') as f:
            return pickle.load(f)
    # Sinon charger depuis Excel
    elif os.path.exists(FICHIER_HT):
        return pd.read_excel(FICHIER_HT)
    else:
        st.error(f"""
        ❌ **Fichier HT introuvable !**
        
        Veuillez placer le fichier `{FICHIER_HT}` dans le même dossier que l'application.
        """)
        st.stop()

def save_data(df, type_tension):
    file = SAVE_FILE_HT if type_tension == 'HT' else SAVE_FILE_BT
    with open(file, 'wb') as f:
        pickle.dump(df, f)

# Initialisation
if 'df_bt' not in st.session_state:
    st.session_state.df_bt = load_data_bt()
if 'df_ht' not in st.session_state:
    st.session_state.df_ht = load_data_ht()

# Header
st.markdown("""
<div class="main-header">
    <h1 class="main-title">⚡ Gestion des Factures Électriques</h1>
    <p class="main-subtitle">SOCODECI & CIE - Plateforme de gestion</p>
</div>
""", unsafe_allow_html=True)

# Sidebar - Sélection du type uniquement
with st.sidebar:
    st.markdown("### 📋 Navigation Principale")
    st.markdown("#### ⚡ CIE")
    
    type_tension = st.radio(
        "Type de tension",
        ["🔌 Basse Tension (BT)", "⚡ Haute Tension (HT)"],
        label_visibility="collapsed"
    )
    
    st.markdown("---")
    
    # Affichage selon le type
    if "BT" in type_tension:
        st.markdown("""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 1.5rem; 
                    border-radius: 10px;'>
            <h3 style='color: white; margin: 0; text-align: center;'>🔌 BASSE TENSION</h3>
            <p style='color: white; margin: 0.5rem 0 0 0; text-align: center;'>Module actif</p>
        </div>
        """, unsafe_allow_html=True)
        
        df_actuel = st.session_state.df_bt
        type_code = "BT"
        
        st.markdown("---")
        st.markdown("### 📊 Statistiques BT")
        
        try:
            montants = pd.to_numeric(df_actuel['MONTANT'], errors='coerce').fillna(0)
            total = montants.sum()
            moyenne = montants.mean()
        except:
            total = 0
            moyenne = 0
        
        st.metric("📝 Lignes", len(df_actuel))
        st.metric("💰 Total", f"{total:,.0f} FCFA")
        st.metric("📊 Moyenne", f"{moyenne:,.0f} FCFA")
    
    else:
        st.markdown("""
        <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); 
                    padding: 1.5rem; 
                    border-radius: 10px;'>
            <h3 style='color: white; margin: 0; text-align: center;'>⚡ HAUTE TENSION</h3>
            <p style='color: white; margin: 0.5rem 0 0 0; text-align: center;'>Module actif</p>
        </div>
        """, unsafe_allow_html=True)
        
        df_actuel = st.session_state.df_ht
        type_code = "HT"
        
        st.markdown("---")
        st.markdown("### 📊 Statistiques HT")
        
        try:
            montants = pd.to_numeric(df_actuel['MONTANT'], errors='coerce').fillna(0)
            total = montants.sum()
            moyenne = montants.mean()
        except:
            total = 0
            moyenne = 0
        
        st.metric("📝 Lignes", len(df_actuel))
        st.metric("💰 Total", f"{total:,.0f} FCFA")
        st.metric("📊 Moyenne", f"{moyenne:,.0f} FCFA")
    
    st.markdown("---")
    st.markdown("#### 💧 SOCODECI")
    st.markdown('<span class="badge-coming">À venir</span>', unsafe_allow_html=True)

# Contenu principal - ONGLETS HORIZONTAUX
if "BT" in type_tension:
    # ===== VUE BASSE TENSION =====
    tab1, tab2 = st.tabs(["📊 Tableau de bord", "🔄 Mise à jour montants"])
    
    with tab1:
        # TABLEAU DE BORD BT
        st.markdown("## 📊 Tableau de bord - Basse Tension (BT)")
        st.markdown("---")
        
        # Statistiques en cartes
        col1, col2, col3, col4 = st.columns(4)
        
        try:
            montants = pd.to_numeric(df_actuel['MONTANT'], errors='coerce').fillna(0)
            total = montants.sum()
            moyenne = montants.mean()
        except:
            total = 0
            moyenne = 0
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #667eea; margin: 0;">📝</h3>
                <h2 style="margin: 0.5rem 0;">{len(df_actuel)}</h2>
                <p style="color: #666; margin: 0;">Total lignes</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #667eea; margin: 0;">💰</h3>
                <h2 style="margin: 0.5rem 0;">{total:,.0f}</h2>
                <p style="color: #666; margin: 0;">Total FCFA</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #667eea; margin: 0;">📊</h3>
                <h2 style="margin: 0.5rem 0;">{moyenne:,.0f}</h2>
                <p style="color: #666; margin: 0;">Moyenne FCFA</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            comptes = df_actuel['COMPTE DE CHARGES'].nunique()
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #667eea; margin: 0;">🔢</h3>
                <h2 style="margin: 0.5rem 0;">{comptes}</h2>
                <p style="color: #666; margin: 0;">Comptes uniques</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Tableau
        st.markdown("### 📋 Données CIE BT")
        st.markdown("*Ajoutez, modifiez ou supprimez des lignes directement*")
        
        edited_df = st.data_editor(
            df_actuel,
            use_container_width=True,
            num_rows="dynamic",
            height=400,
            key="editor_bt"
        )
        
        # Actions
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Sauvegarder", type="primary", use_container_width=True, key="save_bt"):
                st.session_state.df_bt = edited_df
                save_data(edited_df, "BT")
                st.success("✅ Sauvegardé!")
                st.rerun()
        
        with col2:
            # Génération du fichier Excel stylé pour BT
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_actuel.to_excel(writer, index=False, sheet_name='CIE_BT', startrow=6)
                
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                
                workbook = writer.book
                worksheet = writer.sheets['CIE_BT']
                
                # Styles
                header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
                subheader_font = Font(name='Arial', size=12, bold=True, color='2a5298')
                title_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                bold_font = Font(name='Arial', size=11, bold=True)
                
                header_fill = PatternFill(start_color='2a5298', end_color='2a5298', fill_type='solid')
                subheader_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                alt_row_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
                
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thick_border = Border(
                    left=Side(style='medium', color='2a5298'),
                    right=Side(style='medium', color='2a5298'),
                    top=Side(style='medium', color='2a5298'),
                    bottom=Side(style='medium', color='2a5298')
                )
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                # En-tête principal
                worksheet.merge_cells('A1:L2')
                header_cell = worksheet['A1']
                header_cell.value = '⚡ COMPAGNIE IVOIRIENNE D\'ÉLECTRICITÉ (CIE)'
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = center_align
                header_cell.border = thick_border
                worksheet.row_dimensions[1].height = 35
                
                # Sous-en-tête
                worksheet.merge_cells('A3:L3')
                subheader_cell = worksheet['A3']
                subheader_cell.value = 'FACTURES BASSE TENSION (BT) - RAPPORT MENSUEL'
                subheader_cell.font = subheader_font
                subheader_cell.alignment = center_align
                worksheet.row_dimensions[3].height = 25
                
                # Informations
                try:
                    montants = pd.to_numeric(df_actuel['MONTANT'], errors='coerce').fillna(0)
                    total = montants.sum()
                except:
                    total = 0
                
                worksheet['A4'] = 'Date d\'édition:'
                worksheet['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
                worksheet['A5'] = 'Nombre de factures:'
                worksheet['B5'] = len(df_actuel)
                worksheet['E4'] = 'Montant total:'
                worksheet['F4'] = f'{total:,.0f} FCFA'
                worksheet['A4'].font = bold_font
                worksheet['E4'].font = bold_font
                
                # En-tête colonnes
                for col_num, column in enumerate(df_actuel.columns, 1):
                    cell = worksheet.cell(row=7, column=col_num)
                    cell.font = title_font
                    cell.fill = subheader_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                worksheet.row_dimensions[7].height = 30
                
                # Données avec lignes alternées
                start_row = 8
                end_row = start_row + len(df_actuel) - 1
                for row_num in range(start_row, end_row + 1):
                    fill = alt_row_fill if (row_num - start_row) % 2 == 1 else PatternFill()
                    for col_num in range(1, len(df_actuel.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.fill = fill
                
                # Signatures
                derniere_ligne = end_row + 4
                worksheet.merge_cells(f'A{derniere_ligne}:L{derniere_ligne}')
                sig_title = worksheet.cell(row=derniere_ligne, column=1)
                sig_title.value = 'SIGNATURES ET APPROBATIONS'
                sig_title.font = Font(name='Arial', size=12, bold=True, color='2a5298')
                sig_title.alignment = center_align
                worksheet.row_dimensions[derniere_ligne].height = 25
                
                derniere_ligne += 2
                for col_letter, title in [('A', 'PRÉPARÉ PAR'), ('D', 'VÉRIFIÉ PAR'), ('G', 'VALIDÉ PAR'), ('J', 'APPROUVÉ PAR')]:
                    cell = worksheet[f'{col_letter}{derniere_ligne}']
                    cell.value = title
                    cell.font = bold_font
                    cell.alignment = center_align
                    end_col = chr(ord(col_letter) + 2)
                    worksheet.merge_cells(f'{col_letter}{derniere_ligne}:{end_col}{derniere_ligne}')
                
                # Rectangles de signature
                start_row_sig = derniere_ligne + 1
                end_row_sig = start_row_sig + 3
                for col_letter in ['A', 'D', 'G', 'J']:
                    end_col = chr(ord(col_letter) + 2)
                    worksheet.merge_cells(f'{col_letter}{start_row_sig}:{end_col}{end_row_sig}')
                    cell = worksheet[f'{col_letter}{start_row_sig}']
                    cell.border = thin_border
                
                # Footer
                footer_row = end_row_sig + 5
                worksheet.merge_cells(f'A{footer_row}:L{footer_row}')
                footer_cell = worksheet[f'A{footer_row}']
                footer_cell.value = f'Document généré automatiquement - {datetime.now().strftime("%d/%m/%Y à %H:%M")} - CIE Basse Tension'
                footer_cell.font = Font(name='Arial', size=8, italic=True, color='666666')
                footer_cell.alignment = center_align
            
            output.seek(0)
            
            st.download_button(
                "📥 Télécharger Excel BT",
                data=output,
                file_name=f"CIE_BT_Rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_bt"
            )
        
        with col3:
            if st.button("🔄 Actualiser", use_container_width=True, key="refresh_bt"):
                st.rerun()
    
    with tab2:
        # MISE À JOUR BT
        st.markdown("## 🔄 Mise à jour des montants - Basse Tension (BT)")
        st.markdown("---")
        
        st.info("""
        📌 **Configuration automatique BT** :
        - ✅ Colonne clé : **reference contrat**
        - ✅ Colonne montant : **Montant facture TTC**
        - ✅ Colonne caract : **caract** (sera ajoutée au libellé)
        
        ℹ️ *Note: Le fichier source contient "Montant" et "Montant facture TTC", nous utilisons "Montant facture TTC"*
        """)
        
        fichier_source = st.file_uploader(
            "Sélectionnez le fichier source CIE BT",
            type=['xlsx', 'xls'],
            key="upload_bt"
        )
        
        if fichier_source:
            try:
                df_source = pd.read_excel(fichier_source)
                
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #667eea, #764ba2); 
                            color: white; 
                            padding: 1rem; 
                            border-radius: 8px; 
                            margin: 1rem 0;'>
                    <strong>✅ Fichier chargé avec succès</strong><br>
                    📊 {len(df_source)} ligne(s) · 📋 {len(df_source.columns)} colonne(s)
                </div>
                """, unsafe_allow_html=True)
                
                # Configuration des colonnes pour BT
                cle_source = "reference contrat"
                montant_source = "Montant facture TTC"  # On prend "Montant facture TTC" et pas "Montant"
                caract_source = "caract"
                cle_principal = "COMPTE DE CHARGES"
                montant_principal = "MONTANT"
                libelle_principal = "LIBELLE COMPLEMENTAIRE"
                
                # Vérifications
                colonnes_manquantes = []
                if cle_source not in df_source.columns:
                    colonnes_manquantes.append(cle_source)
                if montant_source not in df_source.columns:
                    colonnes_manquantes.append(montant_source)
                if caract_source not in df_source.columns:
                    colonnes_manquantes.append(caract_source)
                
                if colonnes_manquantes:
                    st.error(f"❌ Colonnes manquantes : {', '.join(colonnes_manquantes)}")
                    st.info(f"📋 Colonnes disponibles : {', '.join(df_source.columns)}")
                else:
                    # Récupérer la valeur unique de caract
                    valeur_caract = df_source[caract_source].dropna().unique()
                    if len(valeur_caract) > 0:
                        valeur_caract = str(valeur_caract[0])
                        st.success(f"✅ Valeur 'caract' détectée : **{valeur_caract}**")
                    else:
                        valeur_caract = ""
                        st.warning("⚠️ Aucune valeur trouvée dans la colonne 'caract'")
                    
                    # Aperçu
                    with st.expander("👁️ Aperçu du fichier source"):
                        colonnes_a_afficher = [cle_source, montant_source, caract_source]
                        if "Montant" in df_source.columns:
                            colonnes_a_afficher.insert(1, "Montant")
                        st.dataframe(df_source[colonnes_a_afficher].head(10), use_container_width=True)
                        st.caption(f"💡 Nous utilisons 'Montant facture TTC' pour le montant et '{valeur_caract}' sera ajouté aux libellés")
                    
                    st.markdown("---")
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        if st.button("🔄 LANCER LA MISE À JOUR BT", type="primary", use_container_width=True, key="maj_bt"):
                            with st.spinner("⏳ Mise à jour BT en cours..."):
                                import re
                                
                                dict_montants = dict(zip(
                                    df_source[cle_source].astype(str), 
                                    df_source[montant_source]
                                ))
                                
                                nb_maj = 0
                                nb_libelle_maj = 0
                                
                                for idx, row in df_actuel.iterrows():
                                    cle = str(row[cle_principal])
                                    if cle in dict_montants:
                                        # Mise à jour du montant
                                        df_actuel.at[idx, montant_principal] = dict_montants[cle]
                                        nb_maj += 1
                                        
                                        # Mise à jour du libellé (insérer caract après "CIE BT")
                                        if libelle_principal in df_actuel.columns and valeur_caract:
                                            libelle_actuel = str(df_actuel.at[idx, libelle_principal])
                                            
                                            # Chercher "CIE BT" dans le libellé
                                            if "CIE BT" in libelle_actuel:
                                                # Supprimer l'ancienne valeur caract (format XX/XXXX après CIE BT)
                                                nouveau_libelle = re.sub(
                                                    r'(CIE BT)\s+\d{2}/\d{4}',
                                                    r'\1',
                                                    libelle_actuel
                                                )
                                                
                                                # Insérer la nouvelle valeur juste après "CIE BT"
                                                nouveau_libelle = nouveau_libelle.replace(
                                                    "CIE BT",
                                                    f"CIE BT {valeur_caract}",
                                                    1
                                                )
                                                
                                                df_actuel.at[idx, libelle_principal] = nouveau_libelle
                                                nb_libelle_maj += 1
                                
                                st.session_state.df_bt = df_actuel
                                save_data(df_actuel, "BT")
                                
                                # Résultats
                                st.markdown("---")
                                st.markdown(f"""
                                <div style='background: linear-gradient(135deg, #667eea, #764ba2); 
                                            color: white; 
                                            padding: 2rem; 
                                            border-radius: 10px; 
                                            text-align: center;'>
                                    <h2 style='margin: 0;'>✅ MISE À JOUR BT TERMINÉE</h2>
                                    <p style='margin: 0.5rem 0 0 0;'>{datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                                with col_r1:
                                    st.metric("✅ Montants MAJ", nb_maj)
                                with col_r2:
                                    st.metric("📝 Libellés MAJ", nb_libelle_maj)
                                with col_r3:
                                    st.metric("⚠️ Non trouvées", len(df_actuel) - nb_maj)
                                with col_r4:
                                    taux = (nb_maj / len(df_actuel) * 100) if len(df_actuel) > 0 else 0
                                    st.metric("📈 Taux", f"{taux:.1f}%")
                                
                                st.success(f"🎉 {nb_maj} montant(s) et {nb_libelle_maj} libellé(s) mis à jour avec 'caract: {valeur_caract}' !")
                                st.balloons()
            
            except Exception as e:
                st.error(f"❌ Erreur : {str(e)}")
                st.exception(e)

else:
    # ===== VUE HAUTE TENSION =====
    tab1, tab2 = st.tabs(["📊 Tableau de bord", "🔄 Mise à jour montants"])
    
    with tab1:
        # TABLEAU DE BORD HT
        st.markdown("## 📊 Tableau de bord - Haute Tension (HT)")
        st.markdown("---")
        
        # Statistiques en cartes
        col1, col2, col3, col4 = st.columns(4)
        
        try:
            montants = pd.to_numeric(df_actuel['MONTANT'], errors='coerce').fillna(0)
            total = montants.sum()
            moyenne = montants.mean()
        except:
            total = 0
            moyenne = 0
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #f5576c; margin: 0;">📝</h3>
                <h2 style="margin: 0.5rem 0;">{len(df_actuel)}</h2>
                <p style="color: #666; margin: 0;">Total lignes</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #f5576c; margin: 0;">💰</h3>
                <h2 style="margin: 0.5rem 0;">{total:,.0f}</h2>
                <p style="color: #666; margin: 0;">Total FCFA</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #f5576c; margin: 0;">📊</h3>
                <h2 style="margin: 0.5rem 0;">{moyenne:,.0f}</h2>
                <p style="color: #666; margin: 0;">Moyenne FCFA</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            comptes = df_actuel['COMPTE DE CHARGES'].nunique()
            st.markdown(f"""
            <div class="metric-card">
                <h3 style="color: #f5576c; margin: 0;">🔢</h3>
                <h2 style="margin: 0.5rem 0;">{comptes}</h2>
                <p style="color: #666; margin: 0;">Comptes uniques</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Tableau
        st.markdown("### 📋 Données CIE HT")
        st.markdown("*Ajoutez, modifiez ou supprimez des lignes directement*")
        
        edited_df = st.data_editor(
            df_actuel,
            use_container_width=True,
            num_rows="dynamic",
            height=400,
            key="editor_ht"
        )
        
        # Actions
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Sauvegarder", type="primary", use_container_width=True, key="save_ht"):
                st.session_state.df_ht = edited_df
                save_data(edited_df, "HT")
                st.success("✅ Sauvegardé!")
                st.rerun()
        
        with col2:
            # Génération du fichier Excel stylé pour HT
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_actuel.to_excel(writer, index=False, sheet_name='CIE_HT', startrow=6)
                
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                
                workbook = writer.book
                worksheet = writer.sheets['CIE_HT']
                
                # Styles HT (couleur rose/rouge)
                header_font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
                subheader_font = Font(name='Arial', size=12, bold=True, color='f5576c')
                title_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                bold_font = Font(name='Arial', size=11, bold=True)
                
                header_fill = PatternFill(start_color='2a5298', end_color='2a5298', fill_type='solid')
                subheader_fill = PatternFill(start_color='f5576c', end_color='f5576c', fill_type='solid')
                alt_row_fill = PatternFill(start_color='FFF0F3', end_color='FFF0F3', fill_type='solid')
                
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thick_border = Border(
                    left=Side(style='medium', color='2a5298'),
                    right=Side(style='medium', color='2a5298'),
                    top=Side(style='medium', color='2a5298'),
                    bottom=Side(style='medium', color='2a5298')
                )
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                # En-tête principal
                worksheet.merge_cells('A1:L2')
                header_cell = worksheet['A1']
                header_cell.value = '⚡ COMPAGNIE IVOIRIENNE D\'ÉLECTRICITÉ (CIE)'
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = center_align
                header_cell.border = thick_border
                worksheet.row_dimensions[1].height = 35
                
                # Sous-en-tête
                worksheet.merge_cells('A3:L3')
                subheader_cell = worksheet['A3']
                subheader_cell.value = 'FACTURES HAUTE TENSION (HT) - RAPPORT MENSUEL'
                subheader_cell.font = subheader_font
                subheader_cell.alignment = center_align
                worksheet.row_dimensions[3].height = 25
                
                # Informations
                try:
                    montants = pd.to_numeric(df_actuel['MONTANT'], errors='coerce').fillna(0)
                    total = montants.sum()
                except:
                    total = 0
                
                worksheet['A4'] = 'Date d\'édition:'
                worksheet['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
                worksheet['A5'] = 'Nombre de factures:'
                worksheet['B5'] = len(df_actuel)
                worksheet['E4'] = 'Montant total:'
                worksheet['F4'] = f'{total:,.0f} FCFA'
                worksheet['A4'].font = bold_font
                worksheet['E4'].font = bold_font
                
                # En-tête colonnes
                for col_num, column in enumerate(df_actuel.columns, 1):
                    cell = worksheet.cell(row=7, column=col_num)
                    cell.font = title_font
                    cell.fill = subheader_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                worksheet.row_dimensions[7].height = 30
                
                # Données avec lignes alternées
                start_row = 8
                end_row = start_row + len(df_actuel) - 1
                for row_num in range(start_row, end_row + 1):
                    fill = alt_row_fill if (row_num - start_row) % 2 == 1 else PatternFill()
                    for col_num in range(1, len(df_actuel.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.fill = fill
                
                # Signatures
                derniere_ligne = end_row + 4
                worksheet.merge_cells(f'A{derniere_ligne}:L{derniere_ligne}')
                sig_title = worksheet.cell(row=derniere_ligne, column=1)
                sig_title.value = 'SIGNATURES ET APPROBATIONS'
                sig_title.font = Font(name='Arial', size=12, bold=True, color='f5576c')
                sig_title.alignment = center_align
                worksheet.row_dimensions[derniere_ligne].height = 25
                
                derniere_ligne += 2
                for col_letter, title in [('A', 'PRÉPARÉ PAR'), ('D', 'VÉRIFIÉ PAR'), ('G', 'VALIDÉ PAR'), ('J', 'APPROUVÉ PAR')]:
                    cell = worksheet[f'{col_letter}{derniere_ligne}']
                    cell.value = title
                    cell.font = bold_font
                    cell.alignment = center_align
                    end_col = chr(ord(col_letter) + 2)
                    worksheet.merge_cells(f'{col_letter}{derniere_ligne}:{end_col}{derniere_ligne}')
                
                # Rectangles de signature
                start_row_sig = derniere_ligne + 1
                end_row_sig = start_row_sig + 3
                for col_letter in ['A', 'D', 'G', 'J']:
                    end_col = chr(ord(col_letter) + 2)
                    worksheet.merge_cells(f'{col_letter}{start_row_sig}:{end_col}{end_row_sig}')
                    cell = worksheet[f'{col_letter}{start_row_sig}']
                    cell.border = thin_border
                
                # Footer
                footer_row = end_row_sig + 5
                worksheet.merge_cells(f'A{footer_row}:L{footer_row}')
                footer_cell = worksheet[f'A{footer_row}']
                footer_cell.value = f'Document généré automatiquement - {datetime.now().strftime("%d/%m/%Y à %H:%M")} - CIE Haute Tension'
                footer_cell.font = Font(name='Arial', size=8, italic=True, color='666666')
                footer_cell.alignment = center_align
            
            output.seek(0)
            
            st.download_button(
                "📥 Télécharger Excel HT",
                data=output,
                file_name=f"CIE_HT_Rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_ht"
            )
        
        with col3:
            if st.button("🔄 Actualiser", use_container_width=True, key="refresh_ht"):
                st.rerun()
    
    with tab2:
        # MISE À JOUR HT
        st.markdown("## 🔄 Mise à jour des montants - Haute Tension (HT)")
        st.markdown("---")
        
        st.info("""
        📌 **Configuration automatique HT** :
        - ✅ Colonne clé : **refraccord**
        - ✅ Colonne montant : **montfact**
        - ✅ Colonne caract : **caract** (sera ajoutée au libellé)
        
        ℹ️ *Note: Le fichier source contient "Montant" et "montfact", nous utilisons "montfact"*
        """)
        
        fichier_source = st.file_uploader(
            "Sélectionnez le fichier source CIE HT",
            type=['xlsx', 'xls'],
            key="upload_ht"
        )
        
        if fichier_source:
            try:
                df_source = pd.read_excel(fichier_source)
                
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #f093fb, #f5576c); 
                            color: white; 
                            padding: 1rem; 
                            border-radius: 8px; 
                            margin: 1rem 0;'>
                    <strong>✅ Fichier chargé avec succès</strong><br>
                    📊 {len(df_source)} ligne(s) · 📋 {len(df_source.columns)} colonne(s)
                </div>
                """, unsafe_allow_html=True)
                
                # Configuration des colonnes pour HT
                cle_source = "refraccord"
                montant_source = "montfact"  # On prend "montfact" et pas "Montant"
                caract_source = "caract"
                cle_principal = "COMPTE DE CHARGES"
                montant_principal = "MONTANT"
                libelle_principal = "LIBELLE COMPLEMENTAIRE"
                
                # Vérifications
                colonnes_manquantes = []
                if cle_source not in df_source.columns:
                    colonnes_manquantes.append(cle_source)
                if montant_source not in df_source.columns:
                    colonnes_manquantes.append(montant_source)
                if caract_source not in df_source.columns:
                    colonnes_manquantes.append(caract_source)
                
                if colonnes_manquantes:
                    st.error(f"❌ Colonnes manquantes : {', '.join(colonnes_manquantes)}")
                    st.info(f"📋 Colonnes disponibles : {', '.join(df_source.columns)}")
                else:
                    # Récupérer la valeur unique de caract
                    valeur_caract = df_source[caract_source].dropna().unique()
                    if len(valeur_caract) > 0:
                        valeur_caract = str(valeur_caract[0])
                        st.success(f"✅ Valeur 'caract' détectée : **{valeur_caract}**")
                    else:
                        valeur_caract = ""
                        st.warning("⚠️ Aucune valeur trouvée dans la colonne 'caract'")
                    
                    # Aperçu
                    with st.expander("👁️ Aperçu du fichier source"):
                        colonnes_a_afficher = [cle_source, montant_source, caract_source]
                        if "Montant" in df_source.columns:
                            colonnes_a_afficher.insert(1, "Montant")
                        st.dataframe(df_source[colonnes_a_afficher].head(10), use_container_width=True)
                        st.caption(f"💡 Nous utilisons 'montfact' pour le montant et '{valeur_caract}' sera ajouté aux libellés")
                    
                    st.markdown("---")
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        if st.button("🔄 LANCER LA MISE À JOUR HT", type="primary", use_container_width=True, key="maj_ht"):
                            with st.spinner("⏳ Mise à jour HT en cours..."):
                                import re
                                
                                dict_montants = dict(zip(
                                    df_source[cle_source].astype(str), 
                                    df_source[montant_source]
                                ))
                                
                                nb_maj = 0
                                nb_libelle_maj = 0
                                
                                for idx, row in df_actuel.iterrows():
                                    cle = str(row[cle_principal])
                                    if cle in dict_montants:
                                        # Mise à jour du montant
                                        df_actuel.at[idx, montant_principal] = dict_montants[cle]
                                        nb_maj += 1
                                        
                                        # Mise à jour du libellé (insérer caract après "CIE HT")
                                        if libelle_principal in df_actuel.columns and valeur_caract:
                                            libelle_actuel = str(df_actuel.at[idx, libelle_principal])
                                            
                                            # Chercher "CIE HT" dans le libellé
                                            if "CIE HT" in libelle_actuel:
                                                # Supprimer l'ancienne valeur caract (format XX-XXXX après CIE HT)
                                                nouveau_libelle = re.sub(
                                                    r'(CIE HT)\s+\d{2}-\d{4}',
                                                    r'\1',
                                                    libelle_actuel
                                                )
                                                
                                                # Insérer la nouvelle valeur juste après "CIE HT"
                                                nouveau_libelle = nouveau_libelle.replace(
                                                    "CIE HT",
                                                    f"CIE HT {valeur_caract}",
                                                    1
                                                )
                                                
                                                df_actuel.at[idx, libelle_principal] = nouveau_libelle
                                                nb_libelle_maj += 1
                                
                                st.session_state.df_ht = df_actuel
                                save_data(df_actuel, "HT")
                                
                                # Résultats
                                st.markdown("---")
                                st.markdown(f"""
                                <div style='background: linear-gradient(135deg, #f093fb, #f5576c); 
                                            color: white; 
                                            padding: 2rem; 
                                            border-radius: 10px; 
                                            text-align: center;'>
                                    <h2 style='margin: 0;'>✅ MISE À JOUR HT TERMINÉE</h2>
                                    <p style='margin: 0.5rem 0 0 0;'>{datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                                with col_r1:
                                    st.metric("✅ Montants MAJ", nb_maj)
                                with col_r2:
                                    st.metric("📝 Libellés MAJ", nb_libelle_maj)
                                with col_r3:
                                    st.metric("⚠️ Non trouvées", len(df_actuel) - nb_maj)
                                with col_r4:
                                    taux = (nb_maj / len(df_actuel) * 100) if len(df_actuel) > 0 else 0
                                    st.metric("📈 Taux", f"{taux:.1f}%")
                                
                                st.success(f"🎉 {nb_maj} montant(s) et {nb_libelle_maj} libellé(s) mis à jour avec 'caract: {valeur_caract}' !")
                                st.balloons()
            
            except Exception as e:
                st.error(f"❌ Erreur : {str(e)}")
                st.exception(e)

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; padding: 2rem; color: #666;'>
    <p><strong>SOCODECI & CIE</strong> - Version 2.0 - BT & HT Actifs</p>
</div>
""", unsafe_allow_html=True)
